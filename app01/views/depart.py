# -*- coding : utf-8 -*-            
# @Time : 2022/5/1 10:06
# @Author : SXQ
# @FileName : depart
from django.http import HttpResponse
from django.shortcuts import render, redirect
from app01 import models
from openpyxl import load_workbook


def depart_list(request):
    """
    部门列表
    :param request:
    :return:
    """
    queryset = models.Department.objects.all()

    return render(request, "depart_list.html", {'queryset': queryset})


def depart_add(request):
    """
    添加页面
    :param request:
    :return:
    """
    if request.method == "GET":
        return render(request, "depart_add.html")
    # 获取到标题
    title = request.POST.get('title')
    # 保存到数据库
    models.Department.objects.create(title=title)
    # 重定向回去
    return redirect("/depart/list/")


def depart_delete(request):
    """
    删除部门
    :param request:
    :return:
    """
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """
    修改部门
    :param request:
    :return:
    """
    if request.method == "GET":
        # 根据nid获取
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {'row_object': row_object})
    title = request.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


def depart_multi(request):
    """
    批量上传
    :param request:
    :return:
    """
    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    # 对象传递给openpyxl,并读取文件的内容
    wb = load_workbook(file_object)
    # 得到表格中的sheet
    sheet = wb.worksheets[0]
    # 循环获取每一个单元格中的数据
    # cell = sheet.cell(1, 1)
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        print(text)
        # exists = models.Department.objects.filter(title=text).exists()
        # if not exists:
        #     models.Department.objects.create(title=text)
    return redirect("/depart/list/")
